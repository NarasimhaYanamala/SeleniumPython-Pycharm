import openpyxl
from openpyxl.styles import PatternFill

def getRowsCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook [sheetName]
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook [sheetName]
    return(sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rownum,columnno)

def WriteData(file,sheetName,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rownum,columnno).value = data
    workbook.save(file)

def fillGreenColor(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color='60b212',end_color='60b212',fill_type='solid')
    sheet.cell(rownum,columnno).fill=greenFill
    workbook.save(file)

def fillRedColor(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    RedFill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
    sheet.cell(rownum, columnno).fill = RedFill
    workbook.save(file)





