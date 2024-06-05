import openpyxl

from selenium import webdriver

##File--> Workbook-->Sheets-->Rows-->Cells
file = "C:\\softwares\\sampledata.xlsx"
#file = "C:\\softwares\\data.xlsx"
#workBook = openpyxl.load_workbook(file)
WB = openpyxl.load_workbook(file)
sheet = WB["Sheet1"]
rows = sheet.max_row  ## count no of rows 6
cols = sheet.max_column  ## count no of coloumns 4

##reading all rows & Columns from Excel sheet#### (##nested loop one for columns and rows###)
for r in range(1, rows + 1):  ## starting from 1st row to max rows
    for c in range(1, cols + 1):
        print(sheet.cell(r, c).value, end="  ")
    print()
