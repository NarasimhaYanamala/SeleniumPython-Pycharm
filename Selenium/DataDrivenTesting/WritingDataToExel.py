'''' ##notes##
Reading:
-----
data=sheet.cell(r,c).value
writing:
------
sheet.cell(r,c).value = "Welcome"
'''
import openpyxl

#Same Data
''''
file="C:\\Users\\ynrbt\\OneDrive - kwamenkrumahacademy.org\\Automation\\WritedatatoExcel.xlsx"
WB=openpyxl.load_workbook(file) # WB=ork book
sheet=WB["data"]
sheet1=WB.active   # get active keyword from xl if only one sheet

for r in range(1,6):  # 6 means until 5 rows
    for c in range(1, 4):
        sheet.cell(r,c).value="welcome"
WB.save(file)
'''

# Write Different data
file1="C:\\Users\\ynrbt\\OneDrive - kwamenkrumahacademy.org\\Automation\\WriteData-Different.xlsx"
workbook=openpyxl.load_workbook(file1)
sheet1=workbook["data"]

sheet1.cell(1,1).value=123
sheet1.cell(1,2).value="Sam"
sheet1.cell(1,3).value="Manager"

sheet1.cell(2,1).value=124
sheet1.cell(2,2).value="Samuel"
sheet1.cell(2,3).value="Lead"

sheet1.cell(3,1).value=125
sheet1.cell(3,2).value="Sams"
sheet1.cell(3,3).value="Developer"

workbook.save(file1)





